#!/usr/bin/env python3
"""
Fetch today's retail fuel price from EPPO (Energy Policy and Planning Office)
and append it to data/prices.json.

Source: https://www.eppo.go.th/index.php/th/petroleum/price/structure-oil-price
Each business day, EPPO publishes an Excel file with the retail price structure.
This script scrapes the listing page, downloads the latest Excel, extracts
retail prices, and appends to the JSON data file.
"""

import json
import re
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path
from urllib.request import urlopen, Request
from io import BytesIO

# Thai timezone
TH_TZ = timezone(timedelta(hours=7))

EPPO_LIST_URL = "https://www.eppo.go.th/index.php/th/petroleum/price/structure-oil-price"
EPPO_BASE = "https://www.eppo.go.th"
DATA_FILE = Path(__file__).parent.parent / "data" / "prices.json"

# Map Excel row labels (column B) to our JSON keys
# Support both English and Thai labels (EPPO switched to English ~2026)
# Keys are stripped of whitespace for matching
FUEL_MAP = {
    "เบนซิน 95": "gasoline_95",
    "แก๊สโซฮอล์ 95 อี10": "gasohol_95",
    "แก๊สโซฮอล์ 91": "gasohol_91",
    "แก๊สโซฮอล์ 95 อี20": "gasohol_e20",
    "แก๊สโซฮอล์ 95 อี85": "gasohol_e85",
    "ดีเซลหมุนเร็ว": "diesel",
    "ดีเซลหมุนเร็ว บี20": "diesel_b20",
    "ULG95": "gasoline_95",
    "GASOHOL95 E10": "gasohol_95",
    "GASOHOL91": "gasohol_91",
    "GASOHOL95 E20": "gasohol_e20",
    "GASOHOL95 E85": "gasohol_e85",
    "H-DIESEL": "diesel",
    "H-DIESEL B20": "diesel_b20",
}

# Retail price column — auto-detected from header row "RETAIL" label.
# Fallback index if header not found.
RETAIL_PRICE_COL_FALLBACK = 13

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; thai-fuel-price-bot/1.0)"
}


def fetch_page(url: str) -> str:
    req = Request(url, headers=HEADERS)
    with urlopen(req, timeout=30) as resp:
        return resp.read().decode("utf-8")


def fetch_bytes(url: str) -> bytes:
    req = Request(url, headers=HEADERS)
    with urlopen(req, timeout=30) as resp:
        return resp.read()


def find_latest_download_url(html: str) -> str | None:
    """Find the first (latest) download link on the EPPO listing page."""
    pattern = r'href="(/index\.php/th/component/k2/item/download/[^"]+)"'
    match = re.search(pattern, html)
    if match:
        return EPPO_BASE + match.group(1)
    # Also try full URLs
    pattern2 = r'href="(https://www\.eppo\.go\.th/index\.php/th/component/k2/item/download/[^"]+)"'
    match2 = re.search(pattern2, html)
    if match2:
        return match2.group(1)
    return None


def parse_excel(data: bytes) -> dict[str, float]:
    """Parse the EPPO Excel file and extract retail prices."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    ws = wb.active

    prices: dict[str, float] = {}
    retail_col: int | None = None

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue

        # Auto-detect retail price column from header row
        if retail_col is None:
            for i, cell in enumerate(row):
                if cell and str(cell).strip().upper().startswith("RETAIL"):
                    retail_col = i
                    break
            continue  # skip header row itself

        if retail_col is None or len(row) <= retail_col:
            continue

        cell_text = str(row[1]).strip() if row[1] else ""
        # Normalize: collapse multiple spaces for matching
        cell_normalized = " ".join(cell_text.split())

        for label, json_key in FUEL_MAP.items():
            if cell_normalized == label:
                retail_price = row[retail_col]
                if retail_price is not None:
                    try:
                        val = float(retail_price)
                        if val > 5:  # sanity check: retail price must be > 5 THB
                            prices[json_key] = round(val, 2)
                    except (ValueError, TypeError):
                        pass
                break

    wb.close()

    if not prices:
        retail_col = retail_col if retail_col is not None else RETAIL_PRICE_COL_FALLBACK
        print(f"  Retail column detected at index: {retail_col}", file=sys.stderr)

    return prices


def load_data() -> list[dict]:
    if DATA_FILE.exists():
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def save_data(data: list[dict]) -> None:
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
        f.write("\n")


def main():
    today = datetime.now(TH_TZ).strftime("%Y-%m-%d")
    print(f"Fetching EPPO fuel price for {today}...")

    # Load existing data
    data = load_data()

    # Check if today already exists
    existing_dates = {entry["date"] for entry in data}
    if today in existing_dates:
        print(f"Price for {today} already exists, skipping.")
        return

    # Fetch the listing page
    print("Fetching EPPO listing page...")
    html = fetch_page(EPPO_LIST_URL)

    # Find latest download URL
    download_url = find_latest_download_url(html)
    if not download_url:
        print("ERROR: Could not find download link on EPPO page.", file=sys.stderr)
        sys.exit(1)

    print(f"Downloading Excel from: {download_url}")
    excel_data = fetch_bytes(download_url)

    # Parse Excel
    prices = parse_excel(excel_data)
    if not prices:
        # Dump row labels for debugging
        import openpyxl as _xl
        _ws = _xl.load_workbook(BytesIO(excel_data), data_only=True).active
        labels = [str(r[1]).strip() for r in _ws.iter_rows(values_only=True)
                  if r and len(r) > 1 and r[1]]
        print(f"ERROR: Could not extract any prices from Excel.", file=sys.stderr)
        print(f"  Column B labels found: {labels}", file=sys.stderr)
        print(f"  Expected labels: {list(FUEL_MAP.keys())}", file=sys.stderr)
        sys.exit(1)

    print(f"Extracted prices: {prices}")

    # Append to data
    entry = {"date": today, "prices": prices}
    data.append(entry)
    data.sort(key=lambda e: e["date"])

    save_data(data)
    print(f"Saved price for {today} to {DATA_FILE}")


if __name__ == "__main__":
    main()
